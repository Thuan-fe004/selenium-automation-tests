"""
Bài tập Selenium WebDriver - Test Suite hoàn chỉnh với Test Cases và Excel Report
Môi trường: Python 3.x + Selenium 4.x
Cài đặt: 
    pip install selenium openpyxl pandas
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class TestResult:
    """Class để lưu trữ kết quả test case"""
    def __init__(self, test_id, test_name, description, expected, actual, status, screenshot="", notes=""):
        self.test_id = test_id
        self.test_name = test_name
        self.description = description
        self.expected = expected
        self.actual = actual
        self.status = status  # PASSED, FAILED, SKIPPED
        self.execution_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.screenshot = screenshot
        self.notes = notes

class SeleniumTestSuite:
    def __init__(self):
        """Khởi tạo WebDriver với các cấu hình"""
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.wait = WebDriverWait(self.driver, 10)
        self.test_results = []
        self.current_test_suite = ""
        
        print("=" * 80)
        print("SELENIUM WEBDRIVER - BỘ KIỂM THỬ TỰ ĐỘNG")
        print("=" * 80)
        print(f"Thời gian bắt đầu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("✓ WebDriver đã khởi tạo thành công!\n")
    
    def log_result(self, test_id, test_name, description, expected, actual, status, notes=""):
        """Ghi lại kết quả test case"""
        result = TestResult(test_id, test_name, description, expected, actual, status, notes=notes)
        self.test_results.append(result)
        
        # In ra console
        status_symbol = "✓" if status == "PASSED" else "✗"
        print(f"   {status_symbol} {test_id}: {test_name} - {status}")
        if status == "FAILED":
            print(f"      Expected: {expected}")
            print(f"      Actual: {actual}")
            if notes:
                print(f"      Notes: {notes}")
    
    def bai_1_xu_ly_popup(self):
        """
        BÀI 1: Xử lý các loại JavaScript Popup
        Test Cases:
        - TC_01_01: Xử lý JS Alert
        - TC_01_02: Xác minh text của Alert
        - TC_01_03: Xử lý JS Confirm - Accept
        - TC_01_04: Xử lý JS Confirm - Dismiss
        - TC_01_05: Xử lý JS Prompt với input
        - TC_01_06: Xác minh kết quả sau khi nhập Prompt
        """
        print("\n" + "=" * 80)
        print("BÀI 1: XỬ LÝ JAVASCRIPT POPUP")
        print("=" * 80)
        self.current_test_suite = "Bài 1 - Xử lý Popup"
        
        try:
            self.driver.get("https://the-internet.herokuapp.com/javascript_alerts")
            
            # Chờ trang load xong
            self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "h3")))
            time.sleep(1)
            
            # TC_01_01 & TC_01_02: Xử lý Alert
            print("\nTest Case 1: JS Alert")
            try:
                alert_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[text()='Click for JS Alert']"))
                )
                alert_btn.click()
                
                # Chờ alert xuất hiện
                alert = self.wait.until(EC.alert_is_present())
                alert_text = alert.text
                
                self.log_result(
                    "TC_01_01", 
                    "Click button JS Alert",
                    "Click vào button 'Click for JS Alert'",
                    "Alert popup xuất hiện",
                    "Alert popup đã xuất hiện",
                    "PASSED"
                )
                
                expected_text = "I am a JS Alert"
                if alert_text == expected_text:
                    self.log_result(
                        "TC_01_02",
                        "Xác minh text Alert",
                        "Kiểm tra nội dung text của Alert",
                        expected_text,
                        alert_text,
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_01_02",
                        "Xác minh text Alert",
                        "Kiểm tra nội dung text của Alert",
                        expected_text,
                        alert_text,
                        "FAILED"
                    )
                
                alert.accept()
                
                # Xác minh kết quả
                result = self.wait.until(
                    EC.presence_of_element_located((By.ID, "result"))
                ).text
                
                assert "successfully" in result
                
            except Exception as e:
                self.log_result(
                    "TC_01_01",
                    "Xử lý JS Alert",
                    "Xử lý JS Alert popup",
                    "Alert được xử lý thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_01_03 & TC_01_04: Xử lý Confirm
            print("\nTest Case 2: JS Confirm - Dismiss")
            try:
                confirm_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[text()='Click for JS Confirm']"))
                )
                confirm_btn.click()
                
                confirm = self.wait.until(EC.alert_is_present())
                confirm_text = confirm.text
                confirm.dismiss()  # Nhấn Cancel
                
                result = self.wait.until(
                    EC.presence_of_element_located((By.ID, "result"))
                ).text
                
                if "Cancel" in result:
                    self.log_result(
                        "TC_01_03",
                        "JS Confirm - Dismiss",
                        "Click Confirm và nhấn Cancel",
                        "Hiển thị 'You clicked: Cancel'",
                        result,
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_01_03",
                        "JS Confirm - Dismiss",
                        "Click Confirm và nhấn Cancel",
                        "Hiển thị 'You clicked: Cancel'",
                        result,
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_01_03",
                    "JS Confirm - Dismiss",
                    "Xử lý JS Confirm",
                    "Confirm được dismiss thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_01_04: Confirm - Accept
            print("\nTest Case 3: JS Confirm - Accept")
            try:
                confirm_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[text()='Click for JS Confirm']"))
                )
                confirm_btn.click()
                
                confirm = self.wait.until(EC.alert_is_present())
                confirm.accept()  # Nhấn OK
                
                result = self.wait.until(
                    EC.presence_of_element_located((By.ID, "result"))
                ).text
                
                if "Ok" in result:
                    self.log_result(
                        "TC_01_04",
                        "JS Confirm - Accept",
                        "Click Confirm và nhấn OK",
                        "Hiển thị 'You clicked: Ok'",
                        result,
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_01_04",
                        "JS Confirm - Accept",
                        "Click Confirm và nhấn OK",
                        "Hiển thị 'You clicked: Ok'",
                        result,
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_01_04",
                    "JS Confirm - Accept",
                    "Xử lý JS Confirm",
                    "Confirm được accept thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_01_05 & TC_01_06: Xử lý Prompt
            print("\nTest Case 4: JS Prompt với input")
            try:
                prompt_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[text()='Click for JS Prompt']"))
                )
                prompt_btn.click()
                
                prompt = self.wait.until(EC.alert_is_present())
                
                test_input = "Testing Selenium WebDriver"
                prompt.send_keys(test_input)
                prompt.accept()
                
                self.log_result(
                    "TC_01_05",
                    "Nhập text vào Prompt",
                    "Nhập dữ liệu vào JS Prompt",
                    f"Nhập '{test_input}' thành công",
                    "Đã nhập text và accept",
                    "PASSED"
                )
                
                result = self.wait.until(
                    EC.presence_of_element_located((By.ID, "result"))
                ).text
                
                if test_input in result:
                    self.log_result(
                        "TC_01_06",
                        "Xác minh kết quả Prompt",
                        "Kiểm tra text đã nhập hiển thị trong kết quả",
                        f"Kết quả chứa '{test_input}'",
                        result,
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_01_06",
                        "Xác minh kết quả Prompt",
                        "Kiểm tra text đã nhập hiển thị trong kết quả",
                        f"Kết quả chứa '{test_input}'",
                        result,
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_01_05",
                    "Xử lý JS Prompt",
                    "Nhập text vào Prompt",
                    "Prompt được xử lý thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
                
        except Exception as e:
            print(f"\n✗ Lỗi không xử lý được: {str(e)}")
    
    def bai_2_drag_and_drop(self):
        """
        BÀI 2: Kéo và thả (Drag and Drop)
        Test Cases:
        - TC_02_01: Kiểm tra vị trí ban đầu
        - TC_02_02: Thực hiện drag and drop
        - TC_02_03: Xác minh vị trí sau khi kéo thả
        """
        print("\n" + "=" * 80)
        print("BÀI 2: KÉO VÀ THẢ (DRAG AND DROP)")
        print("=" * 80)
        self.current_test_suite = "Bài 2 - Drag and Drop"
        
        try:
            self.driver.get("https://the-internet.herokuapp.com/drag_and_drop")
            
            # Chờ các phần tử load
            element_a = self.wait.until(EC.presence_of_element_located((By.ID, "column-a")))
            element_b = self.wait.until(EC.presence_of_element_located((By.ID, "column-b")))
            
            time.sleep(1)
            
            # TC_02_01: Kiểm tra vị trí ban đầu
            print("\nTest Case 1: Kiểm tra vị trí ban đầu")
            text_a_before = element_a.text
            text_b_before = element_b.text
            
            if text_a_before == "A" and text_b_before == "B":
                self.log_result(
                    "TC_02_01",
                    "Vị trí ban đầu",
                    "Kiểm tra Column A='A' và Column B='B'",
                    "Column A='A', Column B='B'",
                    f"Column A='{text_a_before}', Column B='{text_b_before}'",
                    "PASSED"
                )
            else:
                self.log_result(
                    "TC_02_01",
                    "Vị trí ban đầu",
                    "Kiểm tra Column A='A' và Column B='B'",
                    "Column A='A', Column B='B'",
                    f"Column A='{text_a_before}', Column B='{text_b_before}'",
                    "FAILED"
                )
            
            # TC_02_02: Thực hiện drag and drop
            print("\nTest Case 2: Thực hiện Drag and Drop")
            try:
                actions = ActionChains(self.driver)
                actions.drag_and_drop(element_a, element_b).perform()
                time.sleep(1)
                
                self.log_result(
                    "TC_02_02",
                    "Thực hiện Drag and Drop",
                    "Kéo Column A và thả vào Column B",
                    "Thao tác drag and drop thành công",
                    "Đã thực hiện drag and drop",
                    "PASSED"
                )
                
            except Exception as e:
                self.log_result(
                    "TC_02_02",
                    "Thực hiện Drag and Drop",
                    "Kéo Column A và thả vào Column B",
                    "Thao tác drag and drop thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
                return
            
            # TC_02_03: Xác minh vị trí sau khi kéo thả
            print("\nTest Case 3: Xác minh vị trí sau khi kéo thả")
            text_a_after = element_a.text
            text_b_after = element_b.text
            
            if text_a_after == "B" and text_b_after == "A":
                self.log_result(
                    "TC_02_03",
                    "Xác minh hoán đổi vị trí",
                    "Kiểm tra Column A='B' và Column B='A'",
                    "Column A='B', Column B='A'",
                    f"Column A='{text_a_after}', Column B='{text_b_after}'",
                    "PASSED"
                )
            else:
                self.log_result(
                    "TC_02_03",
                    "Xác minh hoán đổi vị trí",
                    "Kiểm tra Column A='B' và Column B='A'",
                    "Column A='B', Column B='A'",
                    f"Column A='{text_a_after}', Column B='{text_b_after}'",
                    "FAILED"
                )
                
        except Exception as e:
            print(f"\n✗ Lỗi không xử lý được: {str(e)}")
    
    def bai_3_window_handles(self):
        """
        BÀI 3: Chuyển đổi giữa nhiều cửa sổ
        Test Cases:
        - TC_03_01: Lấy window handle của cửa sổ chính
        - TC_03_02: Mở cửa sổ mới
        - TC_03_03: Chuyển sang cửa sổ mới
        - TC_03_04: Xác minh nội dung cửa sổ mới
        - TC_03_05: Đóng cửa sổ mới và quay về cửa sổ chính
        """
        print("\n" + "=" * 80)
        print("BÀI 3: CHUYỂN ĐỔI GIỮA NHIỀU CỬA SỔ")
        print("=" * 80)
        self.current_test_suite = "Bài 3 - Window Handles"
        
        try:
            self.driver.get("https://the-internet.herokuapp.com/windows")
            time.sleep(1)
            
            # TC_03_01: Lấy window handle chính
            print("\nTest Case 1: Lấy window handle chính")
            main_window = self.driver.current_window_handle
            main_title = self.driver.title
            
            if main_window:
                self.log_result(
                    "TC_03_01",
                    "Lấy window handle chính",
                    "Lấy handle của cửa sổ chính",
                    "Handle không rỗng",
                    f"Handle: {main_window[:20]}...",
                    "PASSED"
                )
            else:
                self.log_result(
                    "TC_03_01",
                    "Lấy window handle chính",
                    "Lấy handle của cửa sổ chính",
                    "Handle không rỗng",
                    "Handle rỗng",
                    "FAILED"
                )
            
            # TC_03_02: Mở cửa sổ mới
            print("\nTest Case 2: Mở cửa sổ mới")
            try:
                click_here_link = self.wait.until(
                    EC.element_to_be_clickable((By.LINK_TEXT, "Click Here"))
                )
                click_here_link.click()
                
                # Chờ cửa sổ mới xuất hiện
                self.wait.until(EC.number_of_windows_to_be(2))
                
                all_windows = self.driver.window_handles
                
                if len(all_windows) == 2:
                    self.log_result(
                        "TC_03_02",
                        "Mở cửa sổ mới",
                        "Click 'Click Here' để mở cửa sổ mới",
                        "Có 2 cửa sổ",
                        f"Có {len(all_windows)} cửa sổ",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_03_02",
                        "Mở cửa sổ mới",
                        "Click 'Click Here' để mở cửa sổ mới",
                        "Có 2 cửa sổ",
                        f"Có {len(all_windows)} cửa sổ",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_03_02",
                    "Mở cửa sổ mới",
                    "Click 'Click Here' để mở cửa sổ mới",
                    "Cửa sổ mới được mở",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
                return
            
            # TC_03_03: Chuyển sang cửa sổ mới
            print("\nTest Case 3: Chuyển sang cửa sổ mới")
            try:
                for window in all_windows:
                    if window != main_window:
                        self.driver.switch_to.window(window)
                        new_window = window
                        break
                
                # Chờ cửa sổ mới load
                self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "h3")))
                
                self.log_result(
                    "TC_03_03",
                    "Chuyển sang cửa sổ mới",
                    "Switch to cửa sổ mới",
                    "Chuyển sang cửa sổ mới thành công",
                    f"Đã chuyển sang window: {new_window[:20]}...",
                    "PASSED"
                )
                
            except Exception as e:
                self.log_result(
                    "TC_03_03",
                    "Chuyển sang cửa sổ mới",
                    "Switch to cửa sổ mới",
                    "Chuyển sang cửa sổ mới thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_03_04: Xác minh nội dung
            print("\nTest Case 4: Xác minh nội dung cửa sổ mới")
            try:
                new_title = self.driver.title
                new_content = self.driver.find_element(By.TAG_NAME, "h3").text
                
                if "New Window" in new_title:
                    self.log_result(
                        "TC_03_04",
                        "Xác minh nội dung cửa sổ mới",
                        "Kiểm tra title và content",
                        "Title chứa 'New Window'",
                        f"Title: '{new_title}', Content: '{new_content}'",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_03_04",
                        "Xác minh nội dung cửa sổ mới",
                        "Kiểm tra title và content",
                        "Title chứa 'New Window'",
                        f"Title: '{new_title}'",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_03_04",
                    "Xác minh nội dung cửa sổ mới",
                    "Kiểm tra title và content",
                    "Lấy được title và content",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_03_05: Đóng và quay về cửa sổ chính
            print("\nTest Case 5: Đóng cửa sổ mới và quay về cửa sổ chính")
            try:
                self.driver.close()
                self.driver.switch_to.window(main_window)
                
                current_title = self.driver.title
                
                if current_title == main_title:
                    self.log_result(
                        "TC_03_05",
                        "Quay về cửa sổ chính",
                        "Đóng cửa sổ mới và switch về cửa sổ chính",
                        f"Title: '{main_title}'",
                        f"Title: '{current_title}'",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_03_05",
                        "Quay về cửa sổ chính",
                        "Đóng cửa sổ mới và switch về cửa sổ chính",
                        f"Title: '{main_title}'",
                        f"Title: '{current_title}'",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_03_05",
                    "Quay về cửa sổ chính",
                    "Đóng cửa sổ mới và switch về cửa sổ chính",
                    "Quay về cửa sổ chính thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
                
        except Exception as e:
            print(f"\n✗ Lỗi không xử lý được: {str(e)}")
    
    def bai_4_iframe(self):
        """
        BÀI 4: Tương tác với iframe
        Test Cases:
        - TC_04_01: Tìm và chuyển vào iframe
        - TC_04_02: Xóa nội dung cũ trong editor
        - TC_04_03: Nhập nội dung mới
        - TC_04_04: Xác minh nội dung đã nhập
        - TC_04_05: Quay về trang chính
        """
        print("\n" + "=" * 80)
        print("BÀI 4: TƯƠNG TÁC VỚI IFRAME")
        print("=" * 80)
        self.current_test_suite = "Bài 4 - iFrame"
        
        try:
            self.driver.get("https://the-internet.herokuapp.com/iframe")
            
            # TC_04_01: Chuyển vào iframe
            print("\nTest Case 1: Tìm và chuyển vào iframe")
            try:
                iframe = self.wait.until(
                    EC.presence_of_element_located((By.ID, "mce_0_ifr"))
                )
                self.driver.switch_to.frame(iframe)
                
                self.log_result(
                    "TC_04_01",
                    "Chuyển vào iframe",
                    "Switch to iframe với id='mce_0_ifr'",
                    "Chuyển vào iframe thành công",
                    "Đã chuyển vào iframe",
                    "PASSED"
                )
                
            except Exception as e:
                self.log_result(
                    "TC_04_01",
                    "Chuyển vào iframe",
                    "Switch to iframe với id='mce_0_ifr'",
                    "Chuyển vào iframe thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
                return
            
            # TC_04_02: Xóa nội dung cũ
            print("\nTest Case 2: Xóa nội dung cũ trong editor")
            try:
                editor = self.wait.until(
                    EC.presence_of_element_located((By.ID, "tinymce"))
                )
                old_content = editor.text
                
                # TinyMCE không hỗ trợ clear(), dùng JavaScript hoặc Ctrl+A + Delete
                # Cách 1: Dùng JavaScript (tốt nhất)
                self.driver.execute_script("arguments[0].innerHTML = '';", editor)
                
                # Hoặc Cách 2: Select all và xóa
                # editor.send_keys(Keys.CONTROL + "a")
                # editor.send_keys(Keys.DELETE)
                
                time.sleep(0.5)
                
                self.log_result(
                    "TC_04_02",
                    "Xóa nội dung cũ",
                    "Clear nội dung trong text editor",
                    "Nội dung được xóa",
                    f"Đã xóa nội dung: '{old_content}'",
                    "PASSED"
                )
                
            except Exception as e:
                self.log_result(
                    "TC_04_02",
                    "Xóa nội dung cũ",
                    "Clear nội dung trong text editor",
                    "Nội dung được xóa",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_04_03: Nhập nội dung mới
            print("\nTest Case 3: Nhập nội dung mới")
            try:
                # Không dùng emoji vì ChromeDriver không hỗ trợ ký tự ngoài BMP
                test_text = "Selenium WebDriver dang kiem thu iframe nay!"
                editor.send_keys(test_text)
                time.sleep(0.5)
                
                self.log_result(
                    "TC_04_03",
                    "Nhập nội dung mới",
                    "Nhập text vào editor",
                    f"Nhập '{test_text}' thành công",
                    "Đã nhập text vào editor",
                    "PASSED"
                )
                
            except Exception as e:
                self.log_result(
                    "TC_04_03",
                    "Nhập nội dung mới",
                    "Nhập text vào editor",
                    "Nhập text thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_04_04: Xác minh nội dung
            print("\nTest Case 4: Xác minh nội dung đã nhập")
            try:
                entered_text = editor.text
                
                if test_text in entered_text:
                    self.log_result(
                        "TC_04_04",
                        "Xác minh nội dung",
                        "Kiểm tra text đã nhập có trong editor",
                        f"Editor chứa '{test_text}'",
                        f"Editor có nội dung: '{entered_text}'",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_04_04",
                        "Xác minh nội dung",
                        "Kiểm tra text đã nhập có trong editor",
                        f"Editor chứa '{test_text}'",
                        f"Editor có nội dung: '{entered_text}'",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_04_04",
                    "Xác minh nội dung",
                    "Kiểm tra text đã nhập có trong editor",
                    "Lấy được nội dung editor",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_04_05: Quay về trang chính
            print("\nTest Case 5: Quay về trang chính")
            try:
                self.driver.switch_to.default_content()
                
                # Xác minh đã ra khỏi iframe
                page_title = self.wait.until(
                    EC.presence_of_element_located((By.TAG_NAME, "h3"))
                ).text
                
                self.log_result(
                    "TC_04_05",
                    "Quay về trang chính",
                    "Switch to default content",
                    "Quay về trang chính thành công",
                    f"Đã quay về trang chính, title: '{page_title}'",
                    "PASSED"
                )
                
            except Exception as e:
                self.log_result(
                    "TC_04_05",
                    "Quay về trang chính",
                    "Switch to default content",
                    "Quay về trang chính thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
                
        except Exception as e:
            print(f"\n✗ Lỗi không xử lý được: {str(e)}")
    
    def bai_5_dynamic_controls(self):
        """
        BÀI 5: Kiểm thử với dữ liệu động
        Test Cases:
        - TC_05_01: Kiểm tra checkbox hiển thị ban đầu
        - TC_05_02: Remove checkbox
        - TC_05_03: Xác minh message sau khi remove
        - TC_05_04: Xác minh checkbox đã biến mất
        - TC_05_05: Add checkbox trở lại
        - TC_05_06: Xác minh checkbox xuất hiện lại
        - TC_05_07: Enable input field
        - TC_05_08: Nhập dữ liệu vào input
        """
        print("\n" + "=" * 80)
        print("BÀI 5: KIỂM THỬ VỚI DỮ LIỆU ĐỘNG")
        print("=" * 80)
        self.current_test_suite = "Bài 5 - Dynamic Controls"
        
        try:
            self.driver.get("https://the-internet.herokuapp.com/dynamic_controls")
            time.sleep(1)
            
            # TC_05_01: Kiểm tra checkbox ban đầu
            print("\nTest Case 1: Kiểm tra checkbox hiển thị ban đầu")
            try:
                checkbox = self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='checkbox']"))
                )
                
                if checkbox.is_displayed():
                    self.log_result(
                        "TC_05_01",
                        "Checkbox hiển thị ban đầu",
                        "Kiểm tra checkbox có hiển thị",
                        "Checkbox hiển thị",
                        "Checkbox đang hiển thị",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_05_01",
                        "Checkbox hiển thị ban đầu",
                        "Kiểm tra checkbox có hiển thị",
                        "Checkbox hiển thị",
                        "Checkbox không hiển thị",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_05_01",
                    "Checkbox hiển thị ban đầu",
                    "Kiểm tra checkbox có hiển thị",
                    "Checkbox hiển thị",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_05_02: Remove checkbox
            print("\nTest Case 2: Remove checkbox")
            try:
                remove_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[text()='Remove']"))
                )
                remove_btn.click()
                
                self.log_result(
                    "TC_05_02",
                    "Click button Remove",
                    "Click vào button 'Remove'",
                    "Button được click thành công",
                    "Đã click button Remove",
                    "PASSED"
                )
                
            except Exception as e:
                self.log_result(
                    "TC_05_02",
                    "Click button Remove",
                    "Click vào button 'Remove'",
                    "Button được click thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_05_03: Xác minh message
            print("\nTest Case 3: Xác minh message sau khi remove")
            try:
                message = self.wait.until(
                    EC.presence_of_element_located((By.ID, "message"))
                )
                message_text = message.text
                
                if "gone" in message_text.lower():
                    self.log_result(
                        "TC_05_03",
                        "Message sau khi remove",
                        "Kiểm tra message hiển thị",
                        "Message chứa 'gone'",
                        f"Message: '{message_text}'",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_05_03",
                        "Message sau khi remove",
                        "Kiểm tra message hiển thị",
                        "Message chứa 'gone'",
                        f"Message: '{message_text}'",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_05_03",
                    "Message sau khi remove",
                    "Kiểm tra message hiển thị",
                    "Message hiển thị",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_05_04: Xác minh checkbox biến mất
            print("\nTest Case 4: Xác minh checkbox đã biến mất")
            try:
                # Chờ checkbox biến mất
                self.wait.until(
                    EC.invisibility_of_element_located((By.CSS_SELECTOR, "input[type='checkbox']"))
                )
                
                self.log_result(
                    "TC_05_04",
                    "Checkbox đã biến mất",
                    "Kiểm tra checkbox không còn hiển thị",
                    "Checkbox không hiển thị",
                    "Checkbox đã biến mất khỏi DOM",
                    "PASSED"
                )
                
            except TimeoutException:
                self.log_result(
                    "TC_05_04",
                    "Checkbox đã biến mất",
                    "Kiểm tra checkbox không còn hiển thị",
                    "Checkbox không hiển thị",
                    "Checkbox vẫn còn hiển thị",
                    "FAILED",
                    "Timeout: Checkbox không biến mất"
                )
            
            # TC_05_05: Add checkbox
            print("\nTest Case 5: Add checkbox trở lại")
            try:
                add_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[text()='Add']"))
                )
                add_btn.click()
                
                self.log_result(
                    "TC_05_05",
                    "Click button Add",
                    "Click vào button 'Add'",
                    "Button được click thành công",
                    "Đã click button Add",
                    "PASSED"
                )
                
            except Exception as e:
                self.log_result(
                    "TC_05_05",
                    "Click button Add",
                    "Click vào button 'Add'",
                    "Button được click thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_05_06: Xác minh checkbox xuất hiện
            print("\nTest Case 6: Xác minh checkbox xuất hiện lại")
            try:
                # Chờ message cập nhật
                self.wait.until(
                    EC.text_to_be_present_in_element((By.ID, "message"), "back")
                )
                
                # Chờ checkbox xuất hiện
                checkbox = self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='checkbox']"))
                )
                
                if checkbox.is_displayed():
                    self.log_result(
                        "TC_05_06",
                        "Checkbox xuất hiện lại",
                        "Kiểm tra checkbox hiển thị sau khi Add",
                        "Checkbox hiển thị",
                        "Checkbox đã xuất hiện lại",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_05_06",
                        "Checkbox xuất hiện lại",
                        "Kiểm tra checkbox hiển thị sau khi Add",
                        "Checkbox hiển thị",
                        "Checkbox không hiển thị",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_05_06",
                    "Checkbox xuất hiện lại",
                    "Kiểm tra checkbox hiển thị sau khi Add",
                    "Checkbox hiển thị",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_05_07: Enable input
            print("\nTest Case 7: Enable input field")
            try:
                input_field = self.driver.find_element(By.CSS_SELECTOR, "input[type='text']")
                is_enabled_before = input_field.is_enabled()
                
                enable_btn = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[text()='Enable']"))
                )
                enable_btn.click()
                
                # Chờ input được enable
                self.wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='text']"))
                )
                
                input_field = self.driver.find_element(By.CSS_SELECTOR, "input[type='text']")
                is_enabled_after = input_field.is_enabled()
                
                if not is_enabled_before and is_enabled_after:
                    self.log_result(
                        "TC_05_07",
                        "Enable input field",
                        "Click 'Enable' để enable input",
                        "Input được enable",
                        f"Before: {is_enabled_before}, After: {is_enabled_after}",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_05_07",
                        "Enable input field",
                        "Click 'Enable' để enable input",
                        "Input được enable",
                        f"Before: {is_enabled_before}, After: {is_enabled_after}",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_05_07",
                    "Enable input field",
                    "Click 'Enable' để enable input",
                    "Input được enable",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
            
            # TC_05_08: Nhập dữ liệu
            print("\nTest Case 8: Nhập dữ liệu vào input")
            try:
                input_field = self.driver.find_element(By.CSS_SELECTOR, "input[type='text']")
                test_text = "Test Dynamic Input 2024"
                input_field.clear()
                input_field.send_keys(test_text)
                
                entered_value = input_field.get_attribute("value")
                
                if entered_value == test_text:
                    self.log_result(
                        "TC_05_08",
                        "Nhập dữ liệu vào input",
                        "Nhập text vào input field",
                        f"Input chứa '{test_text}'",
                        f"Input có value: '{entered_value}'",
                        "PASSED"
                    )
                else:
                    self.log_result(
                        "TC_05_08",
                        "Nhập dữ liệu vào input",
                        "Nhập text vào input field",
                        f"Input chứa '{test_text}'",
                        f"Input có value: '{entered_value}'",
                        "FAILED"
                    )
                    
            except Exception as e:
                self.log_result(
                    "TC_05_08",
                    "Nhập dữ liệu vào input",
                    "Nhập text vào input field",
                    "Nhập text thành công",
                    f"Lỗi: {str(e)}",
                    "FAILED",
                    str(e)
                )
                
        except Exception as e:
            print(f"\n✗ Lỗi không xử lý được: {str(e)}")
    
    def export_to_excel(self, filename="Selenium_Test_Results.xlsx"):
        """Xuất kết quả test ra file Excel với format đẹp"""
        print("\n" + "=" * 80)
        print("XUẤT KẾT QUẢ RA FILE EXCEL")
        print("=" * 80)
        
        # Tạo DataFrame
        data = []
        for result in self.test_results:
            data.append({
                'Test ID': result.test_id,
                'Test Name': result.test_name,
                'Description': result.description,
                'Expected Result': result.expected,
                'Actual Result': result.actual,
                'Status': result.status,
                'Execution Time': result.execution_time,
                'Notes': result.notes
            })
        
        df = pd.DataFrame(data)
        
        # Tạo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        
        # Định dạng header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Thêm header
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Thêm dữ liệu
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Tô màu theo status
                if col_num == 6:  # Cột Status
                    if value == "PASSED":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100", bold=True)
                    elif value == "FAILED":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006", bold=True)
        
        # Điều chỉnh độ rộng cột
        column_widths = {
            'A': 12,  # Test ID
            'B': 30,  # Test Name
            'C': 40,  # Description
            'D': 35,  # Expected Result
            'E': 35,  # Actual Result
            'F': 12,  # Status
            'G': 20,  # Execution Time
            'H': 30   # Notes
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Thêm border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Thêm sheet tổng hợp
        ws_summary = wb.create_sheet("Summary")
        
        total_tests = len(self.test_results)
        passed_tests = len([r for r in self.test_results if r.status == "PASSED"])
        failed_tests = len([r for r in self.test_results if r.status == "FAILED"])
        pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        summary_data = [
            ["TỔNG HỢP KẾT QUẢ KIỂM THỬ", ""],
            ["", ""],
            ["Tổng số Test Cases", total_tests],
            ["Test Cases PASSED", passed_tests],
            ["Test Cases FAILED", failed_tests],
            ["Tỷ lệ Pass (%)", f"{pass_rate:.2f}%"],
            ["", ""],
            ["Thời gian thực hiện", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row_num, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_num, column=1, value=label).font = Font(bold=True, size=12)
            ws_summary.cell(row=row_num, column=2, value=value).font = Font(size=12)
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # Tô màu cho summary
        ws_summary.cell(row=1, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_summary.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF", size=14)
        
        ws_summary.cell(row=4, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws_summary.cell(row=5, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Lưu file
        wb.save(filename)
        print(f"\n✓ Đã xuất kết quả ra file: {filename}")
        print(f"\nTÓM TẮT:")
        print(f"  - Tổng số Test Cases: {total_tests}")
        print(f"  - PASSED: {passed_tests}")
        print(f"  - FAILED: {failed_tests}")
        print(f"  - Tỷ lệ Pass: {pass_rate:.2f}%")
    
    def run_all_tests(self):
        """Chạy tất cả các bài kiểm thử và xuất kết quả"""
        start_time = time.time()
        
        try:
            # Chạy từng bài test
            self.bai_1_xu_ly_popup()
            self.bai_2_drag_and_drop()
            self.bai_3_window_handles()
            self.bai_4_iframe()
            self.bai_5_dynamic_controls()
            
            end_time = time.time()
            execution_time = end_time - start_time
            
            print("\n" + "=" * 80)
            print("✓ TẤT CẢ CÁC BÀI KIỂM THỬ ĐÃ HOÀN THÀNH!")
            print("=" * 80)
            print(f"Tổng thời gian thực hiện: {execution_time:.2f} giây")
            
            # Xuất kết quả ra Excel
            self.export_to_excel()
            
        except Exception as e:
            print(f"\n✗ LỖI NGHIÊM TRỌNG: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            print("\n" + "=" * 80)
            print("Đang đóng trình duyệt...")
            time.sleep(2)
            self.driver.quit()
            print("✓ Đã đóng trình duyệt")
            print("=" * 80)

# =============================================================================
# MAIN - CHẠY TEST SUITE
# =============================================================================
if __name__ == "__main__":
    print("""
    ╔════════════════════════════════════════════════════════════════════╗
    ║         SELENIUM WEBDRIVER - BÀI TẬP KIỂM THỬ TỰ ĐỘNG             ║
    ║                                                                    ║
    ║  Bài 1: Xử lý JavaScript Popup (Alert, Confirm, Prompt)           ║
    ║  Bài 2: Kéo và thả (Drag and Drop)                                ║
    ║  Bài 3: Chuyển đổi giữa nhiều cửa sổ (Window Handles)             ║
    ║  Bài 4: Tương tác với iFrame                                       ║
    ║  Bài 5: Kiểm thử với dữ liệu động (Dynamic Controls)              ║
    ║                                                                    ║
    ║  Kết quả sẽ được xuất ra file Excel với format đẹp!               ║
    ╚════════════════════════════════════════════════════════════════════╝
    """)
    
    test_suite = SeleniumTestSuite()
    test_suite.run_all_tests()
    
    print("\n✓ HOÀN TẤT! Vui lòng kiểm tra file Excel để xem kết quả chi tiết.")